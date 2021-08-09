# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
import pandas as pd
import numpy as np
import openpyxl

def auto_gen(date, menu_array, quantity_array):
    # %%
    file_path = '../output/sales_management.xlsx'
    #データフレームの大きさを決めていたせいでバグってた
    # table_source = np.zeros([10, 3])
    # table_source[:, :] = np.nan
    dataframe = pd.DataFrame(index=[], columns=['商品名', '個数'])

    # %%

    for index, product_name in enumerate(menu_array):
        dataframe.loc[index, '商品名'] = product_name
    for index, quantity in enumerate(quantity_array):
        dataframe.loc[index, '個数'] = quantity

    for product_name, quantity in zip(menu_array, quantity_array):
        dataframe.append({'商品名': product_name, '個数': quantity}, ignore_index=True)

    # %%
    #エクセルファイル指定
    workbook = openpyxl.load_workbook(file_path)
    #ワークシート指定
    sheet = workbook['販売個数']
    #最大行
    maxRow = sheet.max_row + 1

    col_date = 2
    col_menu_name = 6
    col_menu_quantity = 7
    for i in reversed(range(1, maxRow)):
        if sheet.cell(row=i, column=col_menu_name).value != None:
            for product_name, quantity in zip(dataframe['商品名'], dataframe['個数']):
                sheet.cell(i + 1, col_date, value=date)
                sheet.cell(i + 1, col_menu_name, value=product_name)
                sheet.cell(i + 1, col_menu_quantity, value=quantity)
                i = i + 1
            break

    workbook.save(file_path)