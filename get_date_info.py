# To add a new cell, type '# %%'
# To add a new markdown cell, type '# %% [markdown]'
# %%
import datetime
import my_util
import pandas
import openpyxl


# %%
file_path = '../output/sales_management.xlsx'
datetime_date = my_util.get_datetime("8/11")
just_date = datetime_date.date()
just_date

# %%
#dataframe
cols = ['日付', '弁当名', '搬入個数', '販売価格', '販売個数', '売上']
dataframe = pandas.DataFrame(index=[], columns=cols)

#エクセルファイル指定
workbook = openpyxl.load_workbook(file_path)
#ワークシート指定
sheet = workbook['販売個数']
#最大行
maxRow = sheet.max_row + 1

#エクセルのデータをdataframeへ挿入
col_date = 2
col_menu_name = 6
col_menu_quantity = 7
col_sell_value = 8
col_sell_quantity = 18
col_total_sell = 23
for index in range(1, maxRow):
    if sheet.cell(row=index, column=col_date).value == datetime_date:
        dataframe.loc[index, '弁当名'] = sheet.cell(index,
                                                    col_menu_name).value
        dataframe.loc[index, '搬入個数'] = sheet.cell(index,
                                                    col_menu_quantity).value
        dataframe.loc[index, '販売価格'] = sheet.cell(index,
                                                    col_sell_value).value
        index = index + 1
